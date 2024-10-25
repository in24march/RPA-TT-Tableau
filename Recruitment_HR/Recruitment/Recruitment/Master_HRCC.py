import pandas as pd
from openpyxl import Workbook, load_workbook
from RecruitSetting import *
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import os

import RecruitSetting
from RecruitSetting import *



def count_country():
    finder_ins = find_file(RCm_filter)
    excel_file = finder_ins.find_ex_time()
    df = pd.read_excel(excel_file)
    def count_keywords(keyword):
        filter_df = df[df['สถานที่ปฏิบัติงาน'].str.contains(keyword, na=False)]
        count = filter_df.shape[0]
        return count
    count_BKK = count_keywords('กรุงเทพฯ')
    count_NMA = count_keywords('นครราชสีมา')
    print(f'กรุงเทพมีคนสมัคร {count_BKK} คน')
    print(f'นครราชศรีมามีคนสมัคร {count_NMA} คน')
    print(f'total รวมทั้งสองจังหวัด {count_NMA + count_BKK} คน')
    return count_BKK, count_NMA
    
def put_to_temp(c_bkk, c_nma,date_get):
    finder_ins = find_file(master_path)
    excel_file = finder_ins.find_ex_time()
    wb = load_workbook(excel_file)
    ws = wb['02_Recruitment_Performance']
    start_val = None  
    end_val = None
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4, max_row=ws.max_row):
        period_val = row[0].value
        site_val = row[1].value
        App_cell = row[2]
        
        if period_val:
            try:
                start, end = period_val.split('-')
                start = int(start.strip())
                end = int(end.strip())
                print(f'Checking period: {start} to {end}')
                
                if start <= date_get.day <= end:
                    print(f'Current day {date_get.day} is within the period')
                    
                    if site_val == 'BKK':
                        App_cell.value = c_bkk
                        print(f'Setting BKK applications to {c_bkk}')
                    elif site_val == 'NMA':
                        App_cell.value = c_nma
                        print(f'Setting NMA applications to {c_nma}')
                        start_val = start
                        end_val = end
                    
                    # Save and break once the cell is updated
            except ValueError:
                print(f'Invalid period format: {period_val}')
            
    wb.save(excel_file)
    return start_val, end_val
    
def sheet_application_nma(start, end, date_get):
    finder_ins = find_file(master_path)
    excel_file = finder_ins.find_ex_time()
    wb = load_workbook(excel_file)
    print(f'{start} - {end}')
    map_ins = SorcingCH()
    map_ins.sheet()
    celendar_file = find_file(celendar_path)
    excel_calendar = celendar_file.find_ex_time()
    df_calen = pd.read_excel(excel_calendar, sheet_name= 'Sheet1')
    df = pd.read_excel(excel_file, sheet_name= '02_Recruitment_Performance')
    month_obj = date_get.strftime('%b')

    if month_obj in df_calen.columns:
        month_data = df_calen[month_obj].dropna().tolist()
        print(f'Data for {month_obj}: {month_data}')

        totals = {col_name: 0 for _, col_name in map_ins.sheetdata.items()}
        
        for value in month_data:
            print(f'{value}')
            start_run ,end_run = value.split('-')
            start_run = int(start_run.strip())
            if start_run <= start:
                for sheet_name , col_name in map_ins.sheetdata.items():
                    if col_name in df.columns:
                        format_find = f"{start_run}-{end_run}"
                        print(f'date running is : {format_find}')
                        filtered_row = df.loc[(df['Period'].str.contains(format_find)) & (df['Site'] == 'NMA')]
                        
                        
                        if not filtered_row.empty:
                            print(f'col is {col_name} for sheet {sheet_name}')
                            result = filtered_row[col_name].values[0]
                            print(result)
                            
                            totals[col_name] += result
                        else:
                            print(f'No data found for Period {format_find} and Site NMA in sheet {sheet_name}')


        
        # เขียนผลรวมลงในไฟล์ Excel
        for sheet_name, col_name in map_ins.sheetdata.items():
            ws = wb[sheet_name]
            if col_name in totals:
                totals_result = totals[col_name]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col= 1, max_col=ws.max_column):
                    app_cell = row[0]
                    if app_cell.value == 'NMA':
                        count_cell = row[1]
                        count_cell.value = totals_result
                        print(f'Written {totals_result} to {sheet_name} in count column for application NMA')
                        break
            
    wb.save(excel_file)   
        
def sheet_application_bkk(start, end, date_get):
    finder_ins = find_file(master_path)
    excel_file = finder_ins.find_ex_time()
    wb = load_workbook(excel_file)
    print(f'{start} - {end}')
    map_ins = SorcingCH()
    map_ins.sheet()
    celendar_file = find_file(celendar_path)
    excel_calendar = celendar_file.find_ex_time()
    df_calen = pd.read_excel(excel_calendar, sheet_name= 'Sheet1')
    df = pd.read_excel(excel_file, sheet_name= '02_Recruitment_Performance')
    month_obj = date_get.strftime('%b')

    if month_obj in df_calen.columns:
        month_data = df_calen[month_obj].dropna().tolist()
        print(f'Data for {month_obj}: {month_data}')

        totals = {col_name: 0 for _, col_name in map_ins.sheetdata.items()}
        
        for value in month_data:
            print(f'{value}')
            start_run ,end_run = value.split('-')
            start_run = int(start_run.strip())
            if start_run <= start:
                for sheet_name , col_name in map_ins.sheetdata.items():
                    if col_name in df.columns:
                        format_find = f"{start_run}-{end_run}"
                        print(f'date running is : {format_find}')
                        filtered_row = df.loc[(df['Period'].str.contains(format_find)) & (df['Site'] == 'BKK')]
                        
                        
                        if not filtered_row.empty:
                            print(f'col is {col_name} for sheet {sheet_name}')
                            result = filtered_row[col_name].values[0]
                            print(result)
                            
                            totals[col_name] += result
                        else:
                            print(f'No data found for Period {format_find} and Site BKK in sheet {sheet_name}')


        
        # เขียนผลรวมลงในไฟล์ Excel
        for sheet_name, col_name in map_ins.sheetdata.items():
            ws = wb[sheet_name]
            if col_name in totals:
                totals_result = totals[col_name]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col= 1, max_col=ws.max_column):
                    app_cell = row[0]
                    if app_cell.value == 'BKK':
                        count_cell = row[1]
                        count_cell.value = totals_result
                        print(f'Written {totals_result} to {sheet_name} in count column for application BKK')
                        break
            
    wb.save(excel_file)   
    
def date_edit(date_get):
    celendar_file = find_file(celendar_path)
    master_file = find_file(master_path)
    excel_calendar = celendar_file.find_ex_time()
    excel_file =  master_file.find_ex_time()
    wb = load_workbook(excel_file)
    df = pd.read_excel(excel_calendar, sheet_name= 'Sheet1')
    month_obj = date_get.strftime('%b')
    d_range_ins = SorcingCH()
    d_range_ins.date_edit()
    print(month_obj)
    if month_obj in df.columns:
        month_data = df[month_obj].dropna().tolist()
        print(f'Data for {month_obj}: {month_data}')
        ws = wb['02_Recruitment_Performance']
        
        first_val = None
        start1 = None
        end1 = None
        row_idx = 2  # Start from row 2 to skip header

        for idx, value in enumerate(month_data):
            if idx == 0:
                first_val = value
                start1, end1 = value.split('-')
                start1 = int(start1.strip())
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=3, value='BKK')
            row_idx += 1
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=3, value='NMA')
            row_idx += 1
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=3, value='Total')
            row_idx += 1
            
            if value:
                print(f'is {value}')
                start, end = value.split('-')
                start = int(start.strip())
                end = int(end.strip())
                date_obj = date_get
                if start <= date_obj.day <= end:
                    for sheet_name in d_range_ins.date_range:
                        if sheet_name in wb.sheetnames:
                            ws1 = wb[sheet_name]
                            date_range_col = None
                            
                            # Find the column index of 'Date Range' dynamically
                            for col in range(1, ws1.max_column + 1):
                                if ws1.cell(row=1, column=col).value == 'Date Range':
                                    date_range_col = col
                                    break
                            
                            if date_range_col is not None:
                                format = f"{start1}-{end} {month_obj} {date_obj.year}"
                                last_row = ws1.max_row - 1
                                for row in range(2, last_row + 1):
                                    ws1.cell(row=row, column=date_range_col).value = format
                                print(f'Added {format} to Date Range column in sheet {sheet_name}')
                            else:
                                print(f'Error: "Date Range" column not found in sheet {sheet_name}')
                        else:
                            print(f'Error: Sheet {sheet_name} not found in Excel file')
                    
    wb.save(excel_file)
    print('Excel file updated successfully.')
    
def Req_perform():
    finder_file = find_file(User_path)
    find_master_file = find_file(master_path)
    Recent_excel = finder_file.find_ex_time()
    master_file = find_master_file.find_ex_time()
    sheet_name = '02_Recruitment_Performance'
    sheet_name2 = '02_Recruitment_Target'
    columns_to_edit = ['Pre-verify Application by AI', 'First Interview by AI Voice Bot', 'Final Interview by HR', 'Confirm Pre-training by AI Voice Bot', 'Pre-training ', 'Hiring']
    column_target = ['Topic', 'Target']
    # Load the master Excel file
    master_workbook = load_workbook(master_file)
    
    # Print sheet names to verify
    print("Sheet names in master workbook:")
    for name in master_workbook.sheetnames:
        print(name)
    
    # Check if the sheet exists
    if sheet_name not in master_workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in master workbook.")
        return
    
    # master_sheet = master_workbook[sheet_name]
    
    # Update the Excel without changing formulas
    update_excel_without_changing_formulas(Recent_excel, master_file, sheet_name, columns_to_edit)
    recru_target(Recent_excel, master_file, sheet_name2, column_target)


def get_col_index(sheet, col_name):
    """
    Returns the index of a column in the sheet based on the column name.
    """
    col_idx = None
    for cell in sheet[1]:  # Assuming the first row contains column names
        if cell.value == col_name:
            col_idx = cell.column
            break
    if col_idx is None:
        print(f"Column {col_name} not found in sheet.")
    return col_idx

def recru_target(recent_excel_path, master_excel_path, sheet_name, columns_to_edit):
    
    recent_df = pd.read_excel(recent_excel_path, sheet_name=sheet_name)
    
    master_workbook = load_workbook(master_excel_path)
    if sheet_name not in master_workbook.sheetnames:
        print(f"Sheet {sheet_name} not found in master workbook.")
        return
    
    master_sheet = master_workbook[sheet_name]
    column_indices = {}
    for col_name in columns_to_edit:
        col_idx = get_col_index(master_sheet, col_name)
        if col_idx:
            column_indices[col_name] = col_idx
            print(f"{col_name}: {col_idx}")
            
    topic_col_idx = column_indices.get('Topic')
    target_col_idx = column_indices.get('Target')
    
    for master_row in master_sheet.iter_rows(min_row=2, max_row=master_sheet.max_row):
        for col_name, col_idx in column_indices.items():
            topic_cell = master_row[topic_col_idx - 1]
            target_cell = master_row[target_col_idx - 1]
            if topic_cell.value in recent_df['Topic'].values:
                # Update the target cell in master sheet based on matching topic
                target_value = recent_df.loc[recent_df['Topic'] == topic_cell.value, 'Target'].values[0]
                target_cell.value = target_value
    
    # Save the updated master workbook
    master_workbook.save(master_excel_path)
    
def update_excel_without_changing_formulas(recent_excel_path, master_excel_path, sheet_name, columns_to_edit):
    # Load the recent Excel file
    recent_df = pd.read_excel(recent_excel_path, sheet_name=sheet_name)
    
    # Load the master Excel file
    master_workbook = load_workbook(master_excel_path)
    
    # Load the master sheet
    if sheet_name not in master_workbook.sheetnames:
        print(f"Sheet {sheet_name} not found in master workbook.")
        return
    
    master_sheet = master_workbook[sheet_name]

    # Create a dictionary of column names to indexes
    column_indices = {}
    for col_name in columns_to_edit:
        col_idx = get_col_index(master_sheet, col_name)
        if col_idx:
            column_indices[col_name] = col_idx
    
    for index, row in recent_df.iterrows():
        if pd.isna(row['Site']):
            continue
        
        # Find the row in master sheet that matches the current 'Week', 'Period', and 'Site'
        for master_row in master_sheet.iter_rows(min_row=2, max_row=master_sheet.max_row):
            if (master_row[0].value == row['Week'] and
                master_row[1].value == row['Period'] and
                master_row[2].value == row['Site']):
                
                # Update cells in the master sheet
                for col_name, col_index in column_indices.items():
                    if col_name in recent_df.columns:
                        cell = master_sheet.cell(row=master_row[0].row, column=col_index)
                        if cell.data_type == 'n':  # Only update numeric cells
                            cell.value = row[col_name]
                break
    
    # Save the updated master workbook
    master_workbook.save(master_excel_path)
    
def rcm_target():
    finder_file = find_file(master_path)
    excel_master = finder_file.find_ex_time()
    wb = load_workbook(excel_master)
    sheet = {'02_Application' : 'Application', '02_Hiring' : 'Hiring'}
    for sheet_name , col_name in sheet.items():
        df = pd.read_excel(excel_master, sheet_name=sheet_name)
        print(f'{sheet_name} : {col_name}')
        print(df)
        
        filter_total = df.loc[df[col_name] == 'Total']
        print(f'Filtered Total: {filter_total}')
        
        if not filter_total.empty:
            result_total = filter_total['Count'].values[0]
            print(f'{result_total}')
            
            ws = wb['02_Recruitment_Target']
            for row in ws.iter_rows():
                topic = row[0]
                if topic.value == col_name:
                    total_cell = row[1]
                    total_cell.value = result_total 
    
    wb.save(excel_master)
    print('Excel file updated successfully.')

# def collect_chanel(date_get):
#     celendar_file = find_file(celendar_path)
#     chanel_file = find_file(RCm_total)
#     excel_calendar = celendar_file.find_ex_time()
#     excel_chanel = chanel_file.find_ex_time()
#     df = pd.read_excel(excel_calendar, sheet_name= 'Sheet1')
#     wb = load_workbook(excel_chanel)
#     month_obj = date_get.strftime('%b')
#     d_range_ins = SorcingCH()
#     d_range_ins.date_edit()
#     print(month_obj)
#     if month_obj in df.columns:
#         month_data = df[month_obj].dropna().tolist()
#         print(f'Data for {month_obj}: {month_data}')
#         ws = wb['Sheet1']
        

#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
#             if row[0].value == "CHANEL":
#                 chanel_row = row[0].row
#                 break

#         start_col = 3  
#         for idx, value in enumerate(month_data):
#             ws.cell(row=chanel_row, column=start_col + idx, value=f"{value} {month_obj}")

#         # บันทึกไฟล์ Excel หลังจากแก้ไข
#     wb.save(excel_chanel)

# def find_date_column(ws, today):
#     for col in ws.iter_cols(min_row=1, max_row=1, min_col=2, max_col=ws.max_column):
#         header = col[0].value
#         if header:
#             date_range = header.split('-')
#             if len(date_range) == 2:
#                 start_date = int(date_range[0].strip())
#                 end_date = int(date_range[1].split()[0].strip())
#                 if start_date <= today <= end_date:
#                     return col[0].column
#     return None

# def count_chanel():
#     RCM_file = find_file(RCm_total)
#     excel_RCM = RCM_file.find_ex_time()
#     chanel_list = SorcingCH()
#     chanel_list.chanel()
    
#     # เปิดไฟล์ RCM เพื่อเขียนข้อมูล
#     wb = load_workbook(excel_RCM)
#     ws = wb['Sheet1']
    
#     # หาวันที่ปัจจุบัน
#     today = datetime.now().day
    
#     # หาตำแหน่งคอลัมน์ที่ตรงกับช่วงวันที่ปัจจุบัน
#     date_col = find_date_column(ws, today)
#     if date_col is None:
#         print(f"No column found for the date {today}.")
#         return
    
#     # ลูปเพื่อหาไฟล์ในโฟลเดอร์ RCm_filter
#     for file_name in os.listdir(RCm_filter):
#         if file_name.endswith('.xlsx'):
#             file_path = os.path.join(RCm_filter, file_name)
            
#             # ลูปเพื่อนับข้อมูลในแต่ละชีท
#             for idx, chanel in enumerate(chanel_list.list_ch):
#                 try:
#                     df_count = pd.read_excel(file_path, sheet_name=chanel)
#                     data_count = len(df_count)
#                 except Exception as e:
#                     print(f"Error reading {chanel} from {file_name}: {e}")
#                     continue
                
#                 # ค้นหาแถวที่ตรงกับชื่อ chanel
#                 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
#                     if row[0].value == chanel:
#                         ws.cell(row=row[0].row, column=date_col, value=data_count)
#                         break

#     # บันทึกไฟล์ Excel
#     wb.save(excel_RCM)
#     print(f"Data has been written to {excel_RCM} successfully.")

def calculate_sum(df, chanel, site, current_date):
    # เลือกข้อมูลตาม CHANEL และ Site
    df_filtered = df[(df['CHANEL'] == chanel) & (df['Site'] == site)]
    
    if df_filtered.empty:
        return 0
    
    # ดึงชื่อคอลัมน์ช่วงวันที่
    date_ranges = df_filtered.columns[2:]
    
    # คำนวณผลรวมสำหรับวันที่ถึง current_date
    df_filtered = df_filtered[date_ranges].iloc[0]
    
    total_sum = 0
    for date_range in date_ranges:
        start_day, end_day = map(int, date_range.split(' ')[0].split('-'))
        month = date_range.split(' ')[1]
        
        # สร้างวันที่ของช่วงเริ่มต้นและสิ้นสุด
        start_date = datetime(current_date.year, datetime.strptime(month, '%b').month, start_day)
        end_date = datetime(current_date.year, datetime.strptime(month, '%b').month, end_day)
        
        # ตรวจสอบว่าช่วงวันที่ครอบคลุมวันที่ปัจจุบันหรือไม่
        if current_date >= start_date:
            total_sum += df_filtered[date_range]
            print(f"Adding {df_filtered[date_range]} from {date_range}, Total Sum: {total_sum}")

    return total_sum

def count_ch(date_get):
    finder_file = find_file(master_path)
    finder_rcm = find_file(RCm_total)
    excel_master = finder_file.find_ex_time()
    excel_rcm = finder_rcm.find_ex_time()
    chanel_list = SorcingCH()
    chanel_list.chanel()
    wb = load_workbook(excel_master)
    df = pd.read_excel(excel_rcm, sheet_name='Sheet1')
    sheet_total = {'02_Sourcing_Channel_All' : 'Sourcing Channel - All',
                   '02_Sourcing_Channel_BKK' : 'Sourcing Channel - BKK',
                   '02_Sourcing_Channel_NMA' : 'Sourcing Channel - NMA'}
    
    for sheet_name, col_name in sheet_total.items():
        ws = wb[sheet_name]
        
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == col_name:
                suffix = col_name.split('-')[-1].strip()
                
                for cell in col[1:]:
                    if cell.value == 'Total':
                        break
                    if cell.value:
                        result = calculate_sum(df, cell.value, suffix, datetime.now())
                        count_cell = ws.cell(row=cell.row, column=col[0].column + 1)
                        count_cell.value = result
                        
    wb.save(excel_master)
        
    
    
def run_master(date_get):
    count_BKK, count_NMA = count_country()
    date_edit(date_get)
    Req_perform()
    start_val, end_val = put_to_temp(count_BKK, count_NMA,date_get)
    sheet_application_nma(start_val, end_val, date_get)
    sheet_application_bkk(start_val, end_val, date_get)
    # rcm_target()
    count_ch(date_get)
    
if __name__ == '__main__':
    # sheet_application_nma(3, 9, datetime.now())
    # sheet_application_bkk(3, 9, datetime.now())
    # count_ch()
    # run_master(datetime.now())
    Req_perform()
    # rcm_target()