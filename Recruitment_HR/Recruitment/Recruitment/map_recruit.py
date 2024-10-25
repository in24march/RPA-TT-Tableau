from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import RecruitSetting
from RecruitSetting import *

def mapdata():
    # finder_ins = find_file(ori_rec)
    # excel_file = finder_ins.find_excel()
    # wb = load_workbook(excel_file)
    excel_files = [f for f in os.listdir(ori_rec) if f.endswith('.xlsx')]
    
    for index, excel_file in enumerate(excel_files, start=1):
        file_path = os.path.join(ori_rec, excel_file)
        wb = load_workbook(file_path)
        sorcing_ins = RecruitSetting.SorcingCH()
        sorcing_ins.chanel()
        sorcing_ins.map()
        header = [cell.value for cell in wb['Sheet1'][1]]
    
        for sheet_name in sorcing_ins.list_ch:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(header)
                print(f"Create sheet: {sheet_name}")
                
        for row in wb['Sheet1'].iter_rows(min_row=2 ,values_only=True):                                
            value = row[22]
            if value in sorcing_ins.s_map:
                target_sheet = wb[sorcing_ins.s_map[value]]
            else:
                target_sheet = wb['Other']
            target_sheet.append(row)
        
        save_path = os.path.join(RCm_filter, f'Recruitment Filter{index}.xlsx')
        wb.save(save_path)

def collect_chanel(date_get):
    celendar_file = find_file(celendar_path)
    chanel_file = find_file(RCm_total)
    excel_calendar = celendar_file.find_ex_time()
    excel_chanel = chanel_file.find_ex_time()
    df = pd.read_excel(excel_calendar, sheet_name= 'Sheet1')
    wb = load_workbook(excel_chanel)
    month_obj = date_get.strftime('%b')
    d_range_ins = SorcingCH()
    d_range_ins.date_edit()
    print(month_obj)
    if month_obj in df.columns:
        month_data = df[month_obj].dropna().tolist()
        print(f'Data for {month_obj}: {month_data}')
        ws = wb['Sheet1']
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "CHANEL":
                chanel_row = row[0].row
                break
            
        start_col = 3  
        for idx, value in enumerate(month_data):
            ws.cell(row=chanel_row, column=start_col + idx, value=f"{value} {month_obj}")
        # บันทึกไฟล์ Excel หลังจากแก้ไข
    wb.save(excel_chanel)
    

def update_excel_with_counts():
    finder_file = find_file(RCm_total)
    excel_total = finder_file.find_ex_time()
    
    chanel_list = SorcingCH()
    chanel_list.chanel()
    
    # Load the master workbook
    sheet_total = {'02_Sourcing_Channel_All': 'Sourcing Channel - All',
                   '02_Sourcing_Channel_BKK': 'Sourcing Channel - BKK',
                   '02_Sourcing_Channel_NMA': 'Sourcing Channel - NMA'}
    
    # Load the Excel file where data will be updated
    wb2 = load_workbook(excel_total)
    df1 = pd.read_excel(excel_total, sheet_name='Sheet1')  # Assuming 'Sheet1' is the sheet with data
    ws2 = wb2['Sheet1']  # Assuming the data is in 'Sheet1'

    for file_name in os.listdir(RCm_filter):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(RCm_filter, file_name)
            for chanel in chanel_list.list_ch:
                df_count = pd.read_excel(file_path, sheet_name=chanel)
                # Convert 'Apply date' to datetime
                df_count['Apply date'] = pd.to_datetime(df_count['Apply date'], dayfirst=True)
                
                type_case = {
                    'กรุงเทพฯ': 'BKK',
                    'นครราชสีมา': 'NMA',
                    'ทั้งหมด' : 'Total'
                }
                
                for type_name, sheet_key in type_case.items():
                    type_df = df_count[df_count['สถานที่ปฏิบัติงาน'].str.contains(type_name, na=False)]
                    
                    # Filter columns in df1 that have date ranges
                    filtered_columns = [col for col in df1.columns if '-' in col and col.endswith('Aug')]
                    
                    for col in filtered_columns:
                        # Extract start and end days from the column name
                        date_range = col.split(' ')[0]
                        start_day, end_day = map(int, date_range.split('-'))  # Convert to integers
                        
                        # Count rows where Apply date falls within the range
                        count_value = type_df[
                            (type_df['Apply date'].dt.day >= start_day) & 
                            (type_df['Apply date'].dt.day <= end_day)
                        ].shape[0]
                        
                        print(f'{chanel} | {type_name} | {date_range} | {count_value} rows')
                        
                        # Update the values in the Excel2 sheet (wb2)
                        for row in ws2.iter_rows(min_row=2):  # Skip header
                            if row[0].value == chanel and row[1].value == sheet_key:
                                col_index = [i for i, cell in enumerate(ws2[1]) if cell.value == col]
                                if col_index:
                                    col_index = col_index[0] + 1  # +1 because openpyxl uses 1-based index
                                    
                                    # Read the current value
                                    current_value = ws2.cell(row=row[0].row, column=col_index).value
                                    if current_value is None:
                                        current_value = 0
                                    
                                    # Check if we need to update the cell
                                    if current_value == 0 or (type_name == 'Total' and current_value == count_value):
                                        ws2.cell(row=row[0].row, column=col_index, value=count_value)
                                        print(f'Updated {chanel} {sheet_key} for {date_range} with {count_value}')
                                    else:
                                        print(f'Skipped updating {chanel} {sheet_key} for {date_range} as value is {current_value}')
                                    
    wb2.save(excel_total)
    print('Excel file updated successfully.')
    
def calculate_totals():
    finder_file = find_file(RCm_total)
    excel_total = finder_file.find_ex_time()
    
    sheet_name = 'Sheet1'
    df = pd.read_excel(excel_total, sheet_name=sheet_name)

    # อัปเดตยอดรวมในแถว 'All'
    channels = df['CHANEL'].unique()
    for channel in channels:
        # กรองข้อมูลสำหรับช่องทางปัจจุบัน
        subset = df[df['CHANEL'] == channel]
        if 'All' in subset['Site'].values:
            all_row_index = subset[subset['Site'] == 'All'].index[0]
            totals = subset[subset['Site'] != 'All'].sum(numeric_only=True)
            df.loc[all_row_index, df.columns[2:]] = totals
            
    # เขียนข้อมูลที่อัปเดตกลับไปยังชีท Excel
    with pd.ExcelWriter(excel_total, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("Update sum total")

def run_map(date_get):
    mapdata()
    collect_chanel(date_get)
    update_excel_with_counts()
    calculate_totals()
    
if __name__ == '__main__':
    # mapdata()
    # update_excel_with_counts()
    calculate_totals()
    # collect_chanel(datetime.now())