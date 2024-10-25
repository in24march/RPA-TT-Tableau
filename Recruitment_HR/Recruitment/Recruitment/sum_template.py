import pandas as pd
from openpyxl import Workbook, load_workbook
from RecruitSetting import *
from datetime import datetime, timedelta
import os

def SumTemplate(dateint_obj):
    # find_excelfile = [file for file in os.listdir(celendar_path) if file.endswith('.xlsx')]
    # if find_excelfile:
    #     first_file = os.path.join(celendar_path, find_excelfile[0])
    finder_ins = find_file(celendar_path)
    excel_file = finder_ins.find_excel()
    df = pd.read_excel(excel_file, sheet_name='Sheet1')
    wb2 = Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    
    sheet2['A1'] = 'Date'
    sheet2['B1'] = 'Total Application'
    
    # find_excelfile = [file for file in os.listdir(RCm_filter) if file.endswith('.xlsx')]
    # if find_excelfile:
    #     first_file = os.path.join(RCm_filter, find_excelfile[0])
    finder_ins = find_file(RCm_filter)
    excel_file = finder_ins.find_excel()
    df2 = pd.read_excel(excel_file)
    total_count = df2.count().max()
    month_now = dateint_obj.strftime('%b')
    # date_now = datetime.now()
    if month_now in df.columns:
        data_to_write = df[month_now].tolist()
        
        # เขียนข้อมูลลงใน Excel ไฟล์ที่สอง
        for idx, value in enumerate(data_to_write, start=2):
            if pd.notna(value):  # ตรวจสอบว่า value ไม่ใช่ NaN
                date_range = value.split('-')
                start_date = int(date_range[0].strip())
                end_date = int(date_range[1].strip())
                
                if start_date <= dateint_obj.day <= end_date:
                    sheet2.cell(row=idx, column=2, value= total_count)
                
                sheet2.cell(row=idx, column=1, value=f'{value} {month_now}')
                
    wb2.save('excel2.xlsx')
    
if __name__ == '__main__':
    SumTemplate(datetime.now() - timedelta(days=1))